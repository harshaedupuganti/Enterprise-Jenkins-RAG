import os
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from backend.report_pipeline import pipeline_config
import logging

logger = logging.getLogger(__name__)

def generate_report(builds: list[dict], failure_analyses: dict) -> str:
    try:
        wb = openpyxl.Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A2744", end_color="1A2744", fill_type="solid")
        
        fill_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        fill_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        fill_yellow = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        fill_orange = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")

        # Sheet 1: Executive Summary
        ws1 = wb.active
        ws1.title = "Executive Summary"
        ws1.append(["KPIs"])
        
        total = len(builds)
        passed = sum(1 for b in builds if b.get("end_result") == "PASS")
        failed = sum(1 for b in builds if b.get("end_result") in ["FAIL", "TIMEOUT", "EXECUTION_ERROR"])
        cbt_issues = sum(1 for k, v in failure_analyses.items() if "CBT" in v.get("category", ""))
        
        ws1.append(["Total Builds", total])
        ws1.append(["Passed", passed])
        ws1.append(["Failed", failed])
        ws1.append(["CBT Issues", cbt_issues])
        ws1.append([])
        
        ws1.append(["Project", "Total", "Pass", "Fail", "Warning", "CBT Issues", "Software Issues"])
        proj_data = {}
        for b in builds:
            p = b.get("project", "Unknown")
            if p not in proj_data: proj_data[p] = {"t":0, "p":0, "f":0, "w":0, "cbt":0, "sw":0}
            proj_data[p]["t"] += 1
            res = b.get("end_result")
            bnum = str(b.get("build_number"))
            cat = failure_analyses.get(bnum, {}).get("category", "")
            if res == "PASS": proj_data[p]["p"] += 1
            elif res == "WARNING": proj_data[p]["w"] += 1
            elif res in ["FAIL", "TIMEOUT", "EXECUTION_ERROR"]:
                proj_data[p]["f"] += 1
                if "CBT" in cat: proj_data[p]["cbt"] += 1
                if "Software" in cat: proj_data[p]["sw"] += 1
                
        for p, d in proj_data.items():
            ws1.append([p, d["t"], d["p"], d["f"], d["w"], d["cbt"], d["sw"]])

        # Sheet 2: Build Details
        ws2 = wb.create_sheet(title="Build Details")
        headers_2 = [
            "SNO", "Build Number", "Project", "Customer", "Product", "Domain", "Location", "Region",
            "Mode", "Test Bench", "Branch", "Start Time", "Runtime (min)",
            "Overall Result", "Pass TC", "Fail TC", "Warn TC", "Failure Reason (AI)", "Confidence Score", "Analysis Stage"
        ]
        ws2.append(headers_2)
        
        for idx, build in enumerate(builds, start=1):
            build_num = str(build.get("build_number", ""))
            tcs = build.get("testcases", [])
            pass_tc = sum(1 for tc in tcs if tc.get("result") == "PASS")
            fail_tc = sum(1 for tc in tcs if tc.get("result") in ["FAIL", "TIMEOUT", "EXECUTION_ERROR"])
            warn_tc = sum(1 for tc in tcs if tc.get("result") == "WARNING")
            runtime_sec = build.get("overall_runtime", 0)
            runtime_min = round(float(runtime_sec) / 60.0, 1) if runtime_sec else 0.0
            
            analysis_data = failure_analyses.get(build_num, {})
            category = analysis_data.get("category", "-")
            conf = analysis_data.get("confidence_score", "-")
            stage = analysis_data.get("analysis_stage", "-")
            
            row_data = [
                idx, build_num, build.get("project", ""), build.get("customer", ""), build.get("product", ""),
                build.get("domain", ""), build.get("location", ""), build.get("region", ""),
                build.get("build_type", "Standard"), build.get("test_bench", ""), build.get("integration_branch", ""),
                str(build.get("start_time", "")), runtime_min, build.get("end_result", ""),
                pass_tc, fail_tc, warn_tc, category, conf, stage
            ]
            ws2.append(row_data)
            
            current_row = ws2.max_row
            result = build.get("end_result", "")
            row_fill = None
            if result == "PASS": row_fill = fill_green
            elif result in ["FAIL", "TIMEOUT", "EXECUTION_ERROR"]:
                if "CBT" in category: row_fill = fill_orange
                else: row_fill = fill_red
            elif result == "WARNING": row_fill = fill_yellow
                
            if row_fill:
                for col_num in range(1, len(headers_2) + 1):
                    ws2.cell(row=current_row, column=col_num).fill = row_fill

        # Sheet 3: AI Failure Analysis
        ws3 = wb.create_sheet(title="AI Failure Analysis")
        ws3.append(["Build Number", "Project", "Mode", "Fault Category", "Failed Testcases", "AI Analysis"])
        failed_builds = [b for b in builds if b.get("end_result") in ["FAIL", "TIMEOUT", "EXECUTION_ERROR"]]
        for build in failed_builds:
            build_number = str(build.get("build_number", ""))
            failed_tcs_str = "\n".join(f"- {tc.get('title','')}" for tc in build.get("testcases", []) if tc.get("result") in ["FAIL", "TIMEOUT", "EXECUTION_ERROR"])
            analysis_data = failure_analyses.get(build_number, {})
            ws3.append([
                build_number, build.get("project", ""), build.get("build_type", "Standard"),
                analysis_data.get("category", "Review Required"), failed_tcs_str, analysis_data.get("text", "No analysis")
            ])
            
        # Sheet 4: CBT Issue Register
        ws4 = wb.create_sheet(title="CBT Issue Register")
        ws4.append(["Build Number", "Test Bench", "Sub Category", "Trigger Step", "Evidence Excerpt", "Action Required"])
        for build in failed_builds:
            build_number = str(build.get("build_number", ""))
            analysis_data = failure_analyses.get(build_number, {})
            if analysis_data.get("severity") == "BENCH_DOWN":
                ws4.append([
                    build_number, build.get("test_bench", ""), analysis_data.get("sub_category", ""),
                    analysis_data.get("trigger_step", ""), analysis_data.get("text", ""), "Assign to Maintenance"
                ])

        # Formatting
        for sheet in [ws1, ws2, ws3, ws4]:
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if sheet.title != "Executive Summary":
                sheet.freeze_panes = "A2"
            for col in sheet.columns:
                col_letter = col[0].column_letter
                sheet.column_dimensions[col_letter].width = 25
                for cell in col:
                    if cell.row > 1: cell.alignment = Alignment(vertical="top", wrap_text=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f"CBT_Intelligence_Report_{timestamp}.xlsx"
        file_path = os.path.join(pipeline_config.OUTPUT_DIR, filename)
        os.makedirs(pipeline_config.OUTPUT_DIR, exist_ok=True)
        wb.save(file_path)
        return file_path
    except Exception as e:
        logger.error(f"[pipeline_excel] Error generating excel: {e}", exc_info=True)
        return None